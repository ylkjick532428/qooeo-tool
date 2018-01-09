# -*- coding: utf-8 -*-
import os
import sys
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string

#只支持xlsx格式
#https://openpyxl.readthedocs.org/en/latest/usage.html#read-an-existing-workbook
#===============================================================================
# ExcelUtility
#===============================================================================
class ExcelUtility(object):
    def __init__(self):
        pass
    
    #===========================================================================
    # get_fileds : 获得属性名数组
    #===========================================================================
    def get_fileds(self, sheet_ranges):
        row_index = "1"
        fields = []
        fields_keys = {}
        for i in range(1, sheet_ranges.max_column + 1, 1):
            tmp_cel_index = get_column_letter(i)
            tmp_index = "%s%s" % (tmp_cel_index, row_index)
            value = sheet_ranges[tmp_index].value
            value = value.lower()
            if value == "":
                print ("error", tmp_index, "is empty")
                
            if value not in fields_keys:
                fields_keys[value] = ""
            else:
                print ("field name repetition, please change it", tmp_index, value)
                sys.exit(0)
            fields.append(value)
        self.fields = fields
    
    #===========================================================================
    # covent_dict :转换
    #===========================================================================
    def covent_dict(self, sheet_ranges, fields):
        row_index = 1
        rows = []
        for row in sheet_ranges.iter_rows():
            tmp_row = {}
            for cel in range(0, len(fields), 1):
                tmp_cel_index = get_column_letter(cel+1)
                tmp_index = "%s%s" % (tmp_cel_index, row_index)
                key = fields[cel]
                value = sheet_ranges[tmp_index].value
                tmp_row[key] = value
            row_index += 1
            rows.append(tmp_row)
        return rows[1:]
    
    #===========================================================================
    # read : xlsx行列文件转换为[{}]数据
    #===========================================================================
    def read(self, file_name, sheet_name=""):
        wb = load_workbook(file_name)
        if sheet_name:
            sheet_ranges = wb[sheet_name]
        else:
            sheet_ranges = wb.active
        self.get_fileds(sheet_ranges)
        rows = self.covent_dict(sheet_ranges, self.fields)
        return rows
    
    def read_rows(self, file_name, sheet_name=""):
        wb = load_workbook(file_name)
        if sheet_name:
            sheet_ranges = wb[sheet_name]
        else:
            sheet_ranges = wb.active
        row_index = 1
        rows = []
        for row in sheet_ranges.iter_rows():
            tmp_row = []
            for cel in range(1, sheet_ranges.max_column + 1, 1):
                tmp_cel_index = get_column_letter(cel+1)
                tmp_index = "%s%s" % (tmp_cel_index, row_index)
                value = sheet_ranges[tmp_index].value
#                 if tmp_index == "P114":
#                     print(value)
#                     sys.exit(0)
                tmp_row.append(str(value))
            row_index += 1
            rows.append(tmp_row)
        return rows
    
        
if __name__ == '__main__':
    excel_utility = ExcelUtility()
#     excel_utility.read("网上办案_嫌疑人次表两抢.xlsx")
    